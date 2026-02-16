#!/usr/bin/env python3
"""
Spreadsheet Manager
===================
Create, read, and modify Excel workbooks with styled headers,
charts, conditional formatting, and auto-fitted columns.

Requires: openpyxl

Usage:
    python spreadsheet_manager.py              # Run demo
    python spreadsheet_manager.py --help       # Show help

Learn more: https://raccoonette.gumroad.com/l/Python-for-automating-every-day-tasks
"""

import os
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter


class SpreadsheetManager:
    """High-level Excel automation: create sheets, style, chart, export."""

    def __init__(self, file_path: str = None):
        if file_path and os.path.exists(file_path):
            self.workbook = load_workbook(file_path)
            self.file_path = file_path
        else:
            self.workbook = Workbook()
            self.file_path = file_path or "output.xlsx"
        self.sheet = self.workbook.active

        # Reusable styles
        self._header_font = Font(bold=True, color="FFFFFF", size=11)
        self._header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        self._header_align = Alignment(horizontal="center", vertical="center")
        self._thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

    # ------------------------------------------------------------------
    # Sheet management
    # ------------------------------------------------------------------

    def create_sheet(self, name: str, data=None, headers=None):
        if name in self.workbook.sheetnames:
            sheet = self.workbook[name]
        else:
            sheet = self.workbook.create_sheet(title=name)

        if headers:
            sheet.append(headers)
            self._style_header_row(sheet, 1, len(headers))

        if data:
            for row in data:
                sheet.append(row)

        self.sheet = sheet
        return sheet

    def select_sheet(self, name: str):
        if name in self.workbook.sheetnames:
            self.sheet = self.workbook[name]
            return self.sheet
        raise ValueError(f"Sheet '{name}' not found")

    # ------------------------------------------------------------------
    # Writing
    # ------------------------------------------------------------------

    def write_headers(self, headers: list, row: int = 1):
        for col, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=row, column=col, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = self._header_align
            cell.border = self._thin_border

    def write_data(self, data: list, start_row: int = 2):
        for r, row_data in enumerate(data, start_row):
            for c, value in enumerate(row_data, 1):
                cell = self.sheet.cell(row=r, column=c, value=value)
                cell.border = self._thin_border

    def add_row(self, values: list):
        self.sheet.append(values)

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------

    def auto_fit_columns(self, sheet=None):
        sheet = sheet or self.sheet
        for col_cells in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    length = len(str(cell.value or ""))
                    if length > max_len:
                        max_len = length
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = max_len + 4

    def format_column_as_currency(self, col: int, start_row: int = 2):
        for row in range(start_row, self.sheet.max_row + 1):
            self.sheet.cell(row=row, column=col).number_format = "$#,##0.00"

    def format_column_as_percent(self, col: int, start_row: int = 2):
        for row in range(start_row, self.sheet.max_row + 1):
            self.sheet.cell(row=row, column=col).number_format = "0.00%"

    def highlight_cells(self, col: int, threshold, color: str = "FFFF00", start_row: int = 2):
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row in range(start_row, self.sheet.max_row + 1):
            cell = self.sheet.cell(row=row, column=col)
            if cell.value is not None and cell.value >= threshold:
                cell.fill = fill

    # ------------------------------------------------------------------
    # Charts
    # ------------------------------------------------------------------

    def add_bar_chart(self, title, data_col, cat_col=1, start_row=1, end_row=None, position="E2"):
        end_row = end_row or self.sheet.max_row
        chart = BarChart()
        chart.title = title
        chart.style = 10
        data = Reference(self.sheet, min_col=data_col, min_row=start_row, max_row=end_row)
        cats = Reference(self.sheet, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 18
        chart.height = 12
        self.sheet.add_chart(chart, position)

    def add_line_chart(self, title, data_col, cat_col=1, start_row=1, end_row=None, position="E2"):
        end_row = end_row or self.sheet.max_row
        chart = LineChart()
        chart.title = title
        chart.style = 10
        data = Reference(self.sheet, min_col=data_col, min_row=start_row, max_row=end_row)
        cats = Reference(self.sheet, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 18
        chart.height = 12
        self.sheet.add_chart(chart, position)

    def add_pie_chart(self, title, data_col, cat_col=1, start_row=1, end_row=None, position="E2"):
        end_row = end_row or self.sheet.max_row
        chart = PieChart()
        chart.title = title
        data = Reference(self.sheet, min_col=data_col, min_row=start_row, max_row=end_row)
        cats = Reference(self.sheet, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 16
        chart.height = 12
        self.sheet.add_chart(chart, position)

    # ------------------------------------------------------------------
    # Save / utility
    # ------------------------------------------------------------------

    def save(self, file_path: str = None):
        path = file_path or self.file_path
        self.workbook.save(path)
        print(f"Saved: {path}")

    def _style_header_row(self, sheet, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = self._header_align
            cell.border = self._thin_border


# ---------------------------------------------------------------------------
# Demo
# ---------------------------------------------------------------------------

def demo():
    """Create a sample sales report to demonstrate all features."""
    mgr = SpreadsheetManager("demo_report.xlsx")

    # Create a sales sheet
    headers = ["Month", "Revenue", "Expenses", "Profit", "Growth"]
    data = [
        ["January", 12500, 8200, 4300, 0.0],
        ["February", 13800, 8500, 5300, 0.10],
        ["March", 15200, 9100, 6100, 0.10],
        ["April", 14100, 8800, 5300, -0.07],
        ["May", 16500, 9600, 6900, 0.17],
        ["June", 18200, 10200, 8000, 0.10],
    ]

    mgr.create_sheet("Sales Report", data=data, headers=headers)
    mgr.format_column_as_currency(2)
    mgr.format_column_as_currency(3)
    mgr.format_column_as_currency(4)
    mgr.format_column_as_percent(5)
    mgr.highlight_cells(col=4, threshold=6000, color="92D050")
    mgr.auto_fit_columns()
    mgr.add_bar_chart("Monthly Revenue", data_col=2, position="G2")
    mgr.add_line_chart("Profit Trend", data_col=4, position="G18")

    # Remove the default empty sheet if it still exists
    if "Sheet" in mgr.workbook.sheetnames and len(mgr.workbook.sheetnames) > 1:
        del mgr.workbook["Sheet"]

    mgr.save()
    print("Demo report created: demo_report.xlsx")


def main():
    if "--help" in sys.argv or "-h" in sys.argv:
        print(__doc__)
        return
    demo()


if __name__ == "__main__":
    main()
